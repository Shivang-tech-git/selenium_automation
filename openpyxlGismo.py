from openpyxl import load_workbook
import allconstants as AC
# ---------------- Extract information from excel template -----------------
wb = load_workbook('Gismo Prefill Template.xlsm')
details = wb['Details']
cov_sh = wb['Coverages']
cont_sh = wb['Contacts']
columns = []
exclusions = []
cov_list = []
cont_list = []
# Currencies, deductions, additional insureds
for col in range(9,details.max_column+1):
    if not details.cell(2,col).value is None:
        field_value=[None]*9
        for row in range(1,9):
            field_value[row] = str(details.cell(row, col).value)
        for ins_row in range(9,details.max_row+1):
            if not details.cell(ins_row,col).value is None:
                field_value.append(str(details.cell(ins_row,col).value))
            elif not details.cell(ins_row + 1, col).value is None:
                field_value.append(str(details.cell(ins_row, col).value))
            else:
                break
        columns.append(field_value)
    else:
        break
# coverages
for cov_col in range(2,cov_sh.max_column+1):
    if not cov_sh.cell(2,cov_col).value is None:
        cov_count = [None]*14
        for cov_row in range(2,14):
            cov_count[cov_row] = str(cov_sh.cell(cov_row,cov_col).value)
        # append additional coverages to cov_count list
        cov_row = 14 # start with 14th row in coverage sheet
        while not cov_sh.cell(cov_row,cov_col).value is None:
            for new_cov in range(0,10):
                cov_count.append(str(cov_sh.cell(cov_row + new_cov,cov_col).value))
            cov_row += 10
        cov_list.append(cov_count)
    else:
        break
# Broker or Insured contact
if cont_sh.cell(3,1).value == 'Broker':
    cont_str = 'Broker'
elif cont_sh.cell(3,1).value == 'Insured':
    cont_str = 'Insured'
# local contacts
for cont_col in range(2,cont_sh.max_column+1):
    if not cont_sh.cell(2,cont_col).value is None:
        cont_count = [None]*7
        # append first contact to list
        for cont_row in range(2,7):
            cont_count[cont_row] = str(cont_sh.cell(cont_row,cont_col).value)
        # append additional contacts to cont_count list
        cont_row = 7 # start with 7th row in contacts sheet
        while not cont_sh.cell(cont_row,cont_col).value is None:
            for new_cont in range(0,3):
                cont_count.append(str(cont_sh.cell(cont_row + new_cont, cont_col).value))
            cont_row += 3
        cont_list.append(cont_count)
    else:
        break
# Exclusions column
for row in range(3,details.max_row):
    if not details.cell(row,AC.exclusions_col).value is None:
        exclusions.append(details.cell(row,AC.exclusions_col).value)

